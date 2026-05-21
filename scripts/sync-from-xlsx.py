#!/usr/bin/env python3
"""
Read prodotti.xlsx (Davide's review file) and write the edits back into
products.json — the build input for scripts/build-catalog.mjs.

Workflow:
    1. Davide edits prodotti.xlsx in Excel and saves.
    2. Run this script: python3 scripts/sync-from-xlsx.py
    3. Run the catalog build: node scripts/build-catalog.mjs
    4. Commit and deploy.

Flags:
    --prune   Also remove products that exist in products.json but NOT
              in prodotti.xlsx. Use when Davide's xlsx is the complete
              catalog (the default mode keeps orphans, in case rows were
              deleted by accident).

Safety:
    - Products are matched by Codice (SKU). Slugs (URL paths) are kept
      stable: changing the Nome in the xlsx does NOT change the slug.
    - Rows present in the xlsx but absent from products.json are added.
    - Without --prune, products in products.json but missing from the
      xlsx are kept (and reported), never silently removed.
    - The script prints a summary before saving so you can sanity-check
      what's about to land.
"""
import argparse
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_FILE = ROOT / 'prodotti.xlsx'
JSON_FILE = ROOT / 'products.json'

SUBCATEGORY_TO_FILTER = {
    'Elmetti': ('Capo', 'elmetti'),
    'Berretti': ('Capo', 'berretti'),
    'Occhiali': ('Capo', 'occhiali'),
    'Udito': ('Capo', 'udito'),
    'Respirazione': ('Capo', 'respirazione'),
    'Anticaduta': ('Corpo', 'anticaduta'),
    'Alta visibilità': ('Corpo', 'alta-visibilita'),
    'Busto': ('Corpo', 'busto'),
    'Multi-protezione': ('Corpo', 'multi-protezione'),
    'Pantaloni': ('Corpo', 'pantaloni'),
    'Rischio chimico': ('Mani', 'rischio-chimico'),
    'Rischio meccanico': ('Mani', 'rischio-meccanico'),
    'Rischio termico': ('Mani', 'rischio-termico'),
    'Scarpe': ('Piedi', 'scarpe'),
    'Stivali': ('Piedi', 'stivali'),
    'Sanitario': ('Sanitario', 'sanitario'),
    'Alberghiero': ('Alberghiero', 'alberghiero'),
}

CATEGORY_ORDER = ['Capo', 'Corpo', 'Mani', 'Piedi', 'Sanitario', 'Alberghiero']


def slugify(s, maxlen=60):
    s = (s or '').lower()
    for a, b in [('à','a'),('á','a'),('â','a'),('ä','a'),
                  ('è','e'),('é','e'),('ê','e'),('ë','e'),
                  ('ì','i'),('í','i'),('î','i'),('ï','i'),
                  ('ò','o'),('ó','o'),('ô','o'),('ö','o'),
                  ('ù','u'),('ú','u'),('û','u'),('ü','u'),
                  ('ç','c')]:
        s = s.replace(a, b)
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    if len(s) > maxlen:
        s = s[:maxlen].rsplit('-', 1)[0]
    return s


def cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        '--prune', action='store_true',
        help='Remove products from products.json that are not present in prodotti.xlsx.',
    )
    args = parser.parse_args()

    if not XLSX_FILE.exists():
        sys.exit(f'Missing {XLSX_FILE.name}. Run build-prodotti-xlsx.py first.')

    wb = load_workbook(XLSX_FILE)
    if 'Prodotti' not in wb.sheetnames:
        sys.exit(f"{XLSX_FILE.name} has no 'Prodotti' sheet.")
    ws = wb['Prodotti']

    # Build header → column index map
    header_cols = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    required = ['Codice', 'Marca', 'Nome', 'Sotto-categoria', 'Descrizione',
                'Perché 1', 'Perché 2', 'Perché 3', 'Immagini',
                'Scheda tecnica', 'Note', '— Fornitore', '— Descrizione orig.']
    missing = [h for h in required if h not in header_cols]
    if missing:
        sys.exit(f'xlsx header missing columns: {missing}')

    existing = json.loads(JSON_FILE.read_text(encoding='utf-8'))
    by_sku = {p['sku'].strip(): p for p in existing}
    seen_skus = set()
    used_slugs = {p['slug'] for p in existing}

    added, updated, unchanged, problems = [], [], [], []

    for r in range(2, ws.max_row + 1):
        sku = cell(ws, r, header_cols['Codice'])
        if not sku:
            continue
        seen_skus.add(sku)

        subcat = cell(ws, r, header_cols['Sotto-categoria'])
        if subcat not in SUBCATEGORY_TO_FILTER:
            problems.append(f'Riga {r} (SKU {sku}): sotto-categoria sconosciuta "{subcat}"')
            continue
        macro, filter_slug = SUBCATEGORY_TO_FILTER[subcat]

        # Assemble fields from the row
        images_raw = cell(ws, r, header_cols['Immagini'])
        images = [line.strip() for line in images_raw.splitlines() if line.strip()]
        bullets = [cell(ws, r, header_cols[f'Perché {i}']) for i in (1, 2, 3)]
        bullets = [b for b in bullets if b]

        name = cell(ws, r, header_cols['Nome'])

        new = {
            'sku':         sku,
            'brand':       cell(ws, r, header_cols['Marca']),
            'name':        name,
            'category':    macro,
            'subcategory': subcat,
            'filterSlug':  filter_slug,
            'description': cell(ws, r, header_cols['Descrizione']),
            'bullets':     bullets,
            'image':       images[0] if images else 'assets/product-placeholder.png',
            'gallery':     images,
            'datasheet':   cell(ws, r, header_cols['Scheda tecnica']),
            '_supplier':   cell(ws, r, header_cols['— Fornitore']),
            '_source_description': cell(ws, r, header_cols['— Descrizione orig.']),
        }

        notes = cell(ws, r, header_cols['Note'])
        if notes:
            new['_notes'] = notes

        if sku in by_sku:
            old = by_sku[sku]
            # Keep slug stable across edits — URLs must not change
            new['slug'] = old['slug']
            changed_fields = []
            for k in new:
                if old.get(k) != new[k]:
                    changed_fields.append(k)
            if changed_fields:
                old.update(new)
                # Remove _notes if it was cleared
                if not notes and '_notes' in old:
                    del old['_notes']
                updated.append((sku, name, changed_fields))
            else:
                unchanged.append(sku)
        else:
            # New product — generate slug, ensure no collision
            base = slugify(name) or slugify(sku)
            slug = base
            n = 2
            while slug in used_slugs:
                slug = f'{base}-{n}'
                n += 1
            used_slugs.add(slug)
            new['slug'] = slug
            existing.append(new)
            added.append((sku, name))

    removed_in_xlsx = [p for p in existing if p['sku'].strip() not in seen_skus]

    if args.prune and removed_in_xlsx:
        keep = {p['sku'].strip() for p in existing} - {p['sku'].strip() for p in removed_in_xlsx}
        existing = [p for p in existing if p['sku'].strip() in keep]

    # Re-sort for stable diffs
    existing.sort(
        key=lambda p: (CATEGORY_ORDER.index(p['category']), p['subcategory'], p['name'].lower())
    )

    # ── Summary ──────────────────────────────────────────────
    print('Sync da prodotti.xlsx → products.json')
    print(f'  Aggiornati:  {len(updated)}')
    print(f'  Aggiunti:    {len(added)}')
    print(f'  Invariati:   {len(unchanged)}')
    pruned_label = 'rimossi (--prune)' if args.prune else 'mantenuti in products.json'
    print(f'  Non in xlsx: {len(removed_in_xlsx)} ({pruned_label})')

    if updated:
        print('\nModifiche (max 20 mostrate):')
        for sku, name, fields in updated[:20]:
            print(f'  {sku:20s}  {name[:50]:50s}  ← {", ".join(fields)}')
        if len(updated) > 20:
            print(f'  ... e altri {len(updated) - 20}')

    if added:
        print('\nNuovi prodotti:')
        for sku, name in added:
            print(f'  + {sku}  {name}')

    if removed_in_xlsx:
        verb = 'Rimossi' if args.prune else 'In products.json ma non in xlsx (controlla se vanno rimossi)'
        print(f'\n{verb}:')
        for p in removed_in_xlsx:
            print(f'  - {p["sku"]}  {p["name"]}')

    if problems:
        print('\nProblemi trovati (queste righe sono state saltate):')
        for msg in problems:
            print(f'  ! {msg}')
        sys.exit('\nRisolvi i problemi e riprova.')

    JSON_FILE.write_text(
        json.dumps(existing, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8',
    )
    print(f'\nScritto {JSON_FILE.name} ({len(existing)} prodotti).')
    print('Ora lancia: node scripts/build-catalog.mjs')


if __name__ == '__main__':
    main()
