#!/usr/bin/env python3
"""
Generate prodotti.xlsx — the review file Davide opens in Excel.

One row per product. Columns are in Italian, formatted for readability:
  - text wrapping on long fields
  - frozen top row
  - dropdowns on Sotto-categoria
  - "Istruzioni" sheet with usage notes
  - reference columns (Fornitore, Descrizione originale) on the right,
    so Davide can see where the data came from without it cluttering
    the editable area

Re-runnable: regenerates prodotti.xlsx from products.json. Davide's edits
should be done IN prodotti.xlsx — running this script would overwrite them.
Use scripts/sync-from-xlsx.py to push his edits back to products.json.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
JSON_FILE = ROOT / 'products.json'
XLSX_FILE = ROOT / 'prodotti.xlsx'

SUBCATEGORIES = [
    'Elmetti', 'Berretti', 'Occhiali', 'Udito', 'Respirazione',
    'Anticaduta', 'Alta visibilità', 'Busto', 'Multi-protezione', 'Pantaloni',
    'Rischio chimico', 'Rischio meccanico', 'Rischio termico',
    'Scarpe', 'Stivali',
    'Sanitario', 'Alberghiero',
]

# Visible columns in order. (header, json_key, width, wrap)
COLUMNS = [
    ('Codice',              'sku',                 14, False),
    ('Marca',               'brand',               16, False),
    ('Nome',                'name',                42, True ),
    ('Sotto-categoria',     'subcategory',         18, False),
    ('Descrizione',         'description',         70, True ),
    ('Perché 1',            'bullet1',             40, True ),
    ('Perché 2',            'bullet2',             40, True ),
    ('Perché 3',            'bullet3',             40, True ),
    ('Immagini',            'images',              40, True ),
    ('Scheda tecnica',      'datasheet',           32, True ),
    ('Note',                'notes',               40, True ),
    ('— Fornitore',         '_supplier',           28, False),
    ('— Descrizione orig.', '_source_description', 50, True ),
]

ISTRUZIONI = [
    ("Come si usa questo file", True),
    ("", False),
    ("Foglio «Prodotti» — una riga per articolo. Sono colonne da compilare:", False),
    ("", False),
    ("• Marca — il marchio del produttore (es. 3M, Delta Plus, U-Power). Se non lo conosci, lascialo vuoto.", False),
    ("• Nome — il nome commerciale del prodotto, come lo chiameresti tu in negozio. Si può cambiare liberamente.", False),
    ("• Sotto-categoria — scegli dal menù a tendina. Determina dove finisce il prodotto nei filtri del sito.", False),
    ("• Descrizione — il testo che appare sotto al nome nella pagina prodotto. 2–4 frasi, voce Davide. Concreto, schietto, vicino, misurato.", False),
    ("• Perché 1 / 2 / 3 — i tre punti elenco della sezione «Perché ci piace». Frasi corte. Niente superlativi.", False),
    ("• Immagini — uno o più URL o nomi file (uno per riga, premi Alt+Invio per andare a capo nella stessa cella). Le foto vanno in assets/products/.", False),
    ("• Scheda tecnica — link al PDF o nome file. Lascia vuoto se non c'è.", False),
    ("• Note — usa questa colonna per lasciare commenti, dubbi, domande. Le leggeremo prima della pubblicazione.", False),
    ("", False),
    ("Le ultime due colonne (Fornitore, Descrizione originale) sono solo riferimento — non modificarle, le useremo per ricontrollare.", False),
    ("", False),
    ("Cosa fare quando hai finito (o quando hai sistemato un po' di righe):", False),
    ("1. Salva il file (Cmd-S o File → Salva).", False),
    ("2. Avvisaci — pubblichiamo noi le modifiche sul sito.", False),
    ("", False),
    ("La colonna «Codice» è il codice articolo dal gestionale. Non cambiarla.", False),
    ("", False),
    ("Voce del marchio — promemoria veloce:", True),
    ("• Concreto: cita il prodotto, la taglia, la norma, il marchio.", False),
    ("• Schietto: frasi corte. Niente «eccellenza», «leader», «soluzioni», «all'avanguardia».", False),
    ("• Vicino: parla al lettore (tu). «Passa a trovarci», non «i clienti possono prendere appuntamento».", False),
    ("• Misurato: prova con il dettaglio, non con il volume. Se non c'è la prova, taglia la frase.", False),
]


def make_xlsx():
    products = json.loads(JSON_FILE.read_text(encoding='utf-8'))

    wb = Workbook()

    # ── Istruzioni sheet ─────────────────────────────────────
    instr = wb.active
    instr.title = 'Istruzioni'
    instr.sheet_view.showGridLines = False
    instr.column_dimensions['A'].width = 110

    for i, (text, is_heading) in enumerate(ISTRUZIONI, start=1):
        cell = instr.cell(row=i, column=1, value=text)
        if is_heading:
            cell.font = Font(name='Arial', size=14, bold=True, color='DE3813')
        else:
            cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        instr.row_dimensions[i].height = 22 if is_heading else 18

    # ── Prodotti sheet ───────────────────────────────────────
    ws = wb.create_sheet('Prodotti')

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1E1E1E')
    header_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    thin = Side(border_style='thin', color='D9D9D9')
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name='Arial', size=11)
    body_align_wrap = Alignment(wrap_text=True, vertical='top')
    body_align_plain = Alignment(vertical='top')
    ref_fill = PatternFill('solid', start_color='F4F4F4')
    ref_font = Font(name='Arial', size=10, color='5D5858', italic=True)

    # Header row
    for col_idx, (header, _key, width, _wrap) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Body rows
    for row_idx, p in enumerate(products, start=2):
        bullets = p.get('bullets') or []
        bullet_vals = [bullets[i] if i < len(bullets) else '' for i in range(3)]
        images = p.get('gallery') or ([p['image']] if p.get('image') and p['image'] not in ('assets/product-placeholder.png', 'assets/item-placeholder.png') else [])

        row_data = {
            'sku':                 p.get('sku', '').strip(),
            'brand':               p.get('brand', ''),
            'name':                p.get('name', ''),
            'subcategory':         p.get('subcategory', ''),
            'description':         p.get('description', ''),
            'bullet1':             bullet_vals[0],
            'bullet2':             bullet_vals[1],
            'bullet3':             bullet_vals[2],
            'images':              '\n'.join(images),
            'datasheet':           p.get('datasheet', ''),
            'notes':               '',
            '_supplier':           p.get('_supplier', ''),
            '_source_description': p.get('_source_description', ''),
        }

        for col_idx, (header, key, _width, wrap) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
            cell.border = body_border
            is_ref = header.startswith('—')
            if is_ref:
                cell.font = ref_font
                cell.fill = ref_fill
            else:
                cell.font = body_font
            cell.alignment = body_align_wrap if wrap else body_align_plain

        # Row height: leave Excel to auto-size, but bump rows with long content
        approx_lines = max(
            1,
            (len(row_data['description']) // 60) + row_data['description'].count('\n') + 1,
            row_data['images'].count('\n') + 1,
        )
        ws.row_dimensions[row_idx].height = min(20 + (approx_lines - 1) * 14, 140)

    # Freeze header row + leftmost two columns (Codice, Marca)
    ws.freeze_panes = 'C2'

    # Auto-filter on header
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    # Dropdown on Sotto-categoria
    subcat_col_idx = next(i for i, (_, k, _, _) in enumerate(COLUMNS, start=1) if k == 'subcategory')
    subcat_letter = get_column_letter(subcat_col_idx)
    dv = DataValidation(
        type='list',
        formula1=f'"{",".join(SUBCATEGORIES)}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle='Sotto-categoria non valida',
        error='Scegli un valore dal menù a tendina.',
    )
    dv.add(f'{subcat_letter}2:{subcat_letter}{ws.max_row}')
    ws.add_data_validation(dv)

    # Default to Prodotti sheet on open
    wb.active = wb.sheetnames.index('Prodotti')

    wb.save(XLSX_FILE)
    print(f'Wrote {XLSX_FILE.name} with {len(products)} products')


if __name__ == '__main__':
    make_xlsx()
